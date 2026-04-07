"""Academic progress service layer — orchestrates DB, processing engine, and file storage."""
from __future__ import annotations

import io
from typing import Optional

import openpyxl
import pandas as pd
from openpyxl.styles import PatternFill
from sqlalchemy import delete, select
from sqlalchemy.orm import Session

from app.models import AdvisingPeriod, DatasetVersion, Major
from app.models.progress_models import AssignmentType, EquivalentCourse, ProgressAssignment
from app.schemas.progress import (
    DataStatus,
    ReportResponse,
    StudentProgressRow,
)
from app.services.dataset_service import dataset_dataframe, get_active_dataset, upload_dataset
from app.services.progress_processing import (
    CELL_COLORS,
    calculate_credits,
    calculate_gpa_for_rows,
    cell_color,
    extract_primary_grade,
    process_progress_report,
)


def _get_major_id(session: Session, major_code: str) -> int:
    major = session.scalar(select(Major).where(Major.code == major_code))
    if major is None:
        raise LookupError(f"Major '{major_code}' not found.")
    return major.id


# ──────────────────────────────────────────────────────────────────
# Equivalent courses
# ──────────────────────────────────────────────────────────────────

def list_equivalents(session: Session, major_code: str) -> list[EquivalentCourse]:
    major_id = _get_major_id(session, major_code)
    return list(session.scalars(
        select(EquivalentCourse).where(EquivalentCourse.major_id == major_id)
        .order_by(EquivalentCourse.alias_code)
    ).all())


def replace_equivalents(session: Session, major_code: str, pairs: list[dict]) -> list[EquivalentCourse]:
    """Bulk-replace all equivalents for a major. Idempotent."""
    major_id = _get_major_id(session, major_code)
    session.execute(delete(EquivalentCourse).where(EquivalentCourse.major_id == major_id))
    new_rows = [
        EquivalentCourse(
            major_id=major_id,
            alias_code=p['alias_code'].strip().upper(),
            canonical_code=p['canonical_code'].strip().upper(),
        )
        for p in pairs
        if p.get('alias_code') and p.get('canonical_code')
    ]
    session.add_all(new_rows)
    session.flush()
    return new_rows


# ──────────────────────────────────────────────────────────────────
# Assignment types (labels: SCE, FEC, …)
# ──────────────────────────────────────────────────────────────────

def list_assignment_types(session: Session, major_code: str) -> list[AssignmentType]:
    major_id = _get_major_id(session, major_code)
    return list(session.scalars(
        select(AssignmentType)
        .where(AssignmentType.major_id == major_id)
        .order_by(AssignmentType.sort_order, AssignmentType.label)
    ).all())


def create_assignment_type(
    session: Session, major_code: str, label: str, sort_order: int = 0
) -> AssignmentType:
    major_id = _get_major_id(session, major_code)
    existing = session.scalar(
        select(AssignmentType).where(
            AssignmentType.major_id == major_id,
            AssignmentType.label == label,
        )
    )
    if existing:
        raise ValueError(f"Assignment type '{label}' already exists.")
    at = AssignmentType(major_id=major_id, label=label, sort_order=sort_order)
    session.add(at)
    session.flush()
    return at


def delete_assignment_type(session: Session, major_code: str, type_id: int) -> None:
    major_id = _get_major_id(session, major_code)
    at = session.scalar(
        select(AssignmentType).where(
            AssignmentType.id == type_id,
            AssignmentType.major_id == major_id,
        )
    )
    if at is None:
        raise LookupError(f"Assignment type {type_id} not found.")
    n = session.query(ProgressAssignment).filter_by(
        major_id=major_id, assignment_type=at.label
    ).count()
    if n > 0:
        raise ValueError(f"Cannot delete '{at.label}': {n} student assignment(s) still use this type.")
    session.delete(at)
    session.flush()


# ──────────────────────────────────────────────────────────────────
# Per-student progress assignments
# ──────────────────────────────────────────────────────────────────

def list_assignments(
    session: Session, major_code: str, student_id: Optional[str] = None
) -> list[ProgressAssignment]:
    major_id = _get_major_id(session, major_code)
    q = select(ProgressAssignment).where(ProgressAssignment.major_id == major_id)
    if student_id:
        q = q.where(ProgressAssignment.student_id == student_id)
    return list(session.scalars(q).all())


def upsert_assignment(
    session: Session,
    major_code: str,
    student_id: str,
    assignment_type: str,
    course_code: str,
) -> ProgressAssignment:
    """Create or update one assignment. Blocks if assignment_type doesn't exist.

    Substitute slot types (ending in '_substitute') are dynamic and bypass the
    registered-type check — they are created automatically for exempted courses.
    """
    major_id = _get_major_id(session, major_code)

    if not assignment_type.endswith('_substitute'):
        at = session.scalar(
            select(AssignmentType).where(
                AssignmentType.major_id == major_id,
                AssignmentType.label == assignment_type,
            )
        )
        if at is None:
            raise LookupError(f"Assignment type '{assignment_type}' does not exist.")

    # Block assignment to a course that's already a required/intensive target
    config = _load_course_config(session, major_code)
    if config:
        all_target = (
            set(config.get('target_courses', {}).keys())
            | set(config.get('intensive_courses', {}).keys())
        )
        if course_code.strip().upper() in all_target:
            raise ValueError(
                f"Course '{course_code}' is already a required/intensive course "
                f"and cannot be assigned as '{assignment_type}'."
            )

    existing = session.scalar(
        select(ProgressAssignment).where(
            ProgressAssignment.major_id == major_id,
            ProgressAssignment.student_id == student_id,
            ProgressAssignment.assignment_type == assignment_type,
        )
    )
    if existing:
        existing.course_code = course_code.strip().upper()
        session.flush()
        return existing

    pa = ProgressAssignment(
        major_id=major_id,
        student_id=student_id,
        assignment_type=assignment_type,
        course_code=course_code.strip().upper(),
    )
    session.add(pa)
    session.flush()
    return pa


def delete_assignment(
    session: Session, major_code: str, student_id: str, assignment_type: str
) -> None:
    major_id = _get_major_id(session, major_code)
    pa = session.scalar(
        select(ProgressAssignment).where(
            ProgressAssignment.major_id == major_id,
            ProgressAssignment.student_id == student_id,
            ProgressAssignment.assignment_type == assignment_type,
        )
    )
    if pa is None:
        raise LookupError("Assignment not found.")
    session.delete(pa)
    session.flush()


def reset_all_assignments(session: Session, major_code: str) -> int:
    """Delete all assignments for a major. Returns count deleted."""
    major_id = _get_major_id(session, major_code)
    n = session.query(ProgressAssignment).filter_by(major_id=major_id).count()
    session.execute(delete(ProgressAssignment).where(ProgressAssignment.major_id == major_id))
    session.flush()
    return n


# ──────────────────────────────────────────────────────────────────
# File uploads (delegate to dataset_service)
# ──────────────────────────────────────────────────────────────────

def upload_progress_report(
    session: Session, major_code: str, filename: str, content: bytes, user_id: int
) -> dict:
    """Parse, validate, and store a progress report. Returns summary."""
    # Validation is done inside _parse_dataset → read_progress_report; raises ValueError on failure.
    version = upload_dataset(
        session,
        major_code=major_code,
        dataset_type='progress_report',
        filename=filename,
        content=content,
        user_id=user_id,
    )
    records = version.parsed_payload.get('records', [])
    ids = {r.get('ID') for r in records if r.get('ID')}
    return {'student_count': len(ids), 'row_count': len(records)}


def preview_progress_upload(session: Session, major_code: str, content: bytes) -> dict:
    """Parse an incoming progress report and diff it against the current active version. No data is saved."""
    from app.services.progress_processing import read_progress_report
    try:
        incoming_df = read_progress_report(content, 'preview.xlsx')
    except ValueError as exc:
        raise ValueError(str(exc)) from exc
    incoming_ids = set(incoming_df['ID'].astype(str).unique()) if 'ID' in incoming_df.columns else set()
    current_df = _load_progress_df(session, major_code)
    if current_df is None:
        return {
            'new_students': len(incoming_ids),
            'removed_students': 0,
            'grade_changes': 0,
            'total_students': len(incoming_ids),
        }
    current_ids = set(current_df['ID'].astype(str).unique())
    new_ids = incoming_ids - current_ids
    removed_ids = current_ids - incoming_ids
    grade_changes = 0
    if 'Grade' in incoming_df.columns and 'Grade' in current_df.columns:
        for sid in incoming_ids & current_ids:
            curr_grades = frozenset(current_df.loc[current_df['ID'] == sid, 'Grade'].astype(str))
            new_grades = frozenset(incoming_df.loc[incoming_df['ID'] == sid, 'Grade'].astype(str))
            if curr_grades != new_grades:
                grade_changes += 1
    return {
        'new_students': len(new_ids),
        'removed_students': len(removed_ids),
        'grade_changes': grade_changes,
        'total_students': len(incoming_ids),
    }


def upload_course_config(
    session: Session, major_code: str, filename: str, content: bytes, user_id: int
) -> dict:
    """Parse, validate, and store a course configuration. Returns summary."""
    version = upload_dataset(
        session,
        major_code=major_code,
        dataset_type='course_config',
        filename=filename,
        content=content,
        user_id=user_id,
    )
    config = version.parsed_payload.get('records', [{}])[0]
    return {
        'required_count': len(config.get('target_courses', {})),
        'intensive_count': len(config.get('intensive_courses', {})),
    }


# ──────────────────────────────────────────────────────────────────
# Internal data loaders
# ──────────────────────────────────────────────────────────────────

def _load_course_config(session: Session, major_code: str) -> dict | None:
    dv = get_active_dataset(session, major_code, 'course_config')
    if dv is None or not dv.parsed_payload:
        return None
    records = dv.parsed_payload.get('records', [])
    return records[0] if records else None


def _load_progress_df(session: Session, major_code: str) -> pd.DataFrame | None:
    dv = get_active_dataset(session, major_code, 'progress_report')
    if dv is None or not dv.parsed_payload:
        return None
    records = dv.parsed_payload.get('records', [])
    if not records:
        return None
    df = pd.DataFrame(records)
    df['ID'] = df['ID'].astype(str).str.strip()
    return df


def _build_equiv_map(session: Session, major_code: str) -> dict[str, str]:
    return {e.alias_code: e.canonical_code for e in list_equivalents(session, major_code)}


def _build_assign_map(session: Session, major_code: str) -> dict[str, dict[str, str]]:
    by_student: dict[str, dict[str, str]] = {}
    for a in list_assignments(session, major_code):
        by_student.setdefault(a.student_id, {})[a.assignment_type] = a.course_code
    return by_student


# ──────────────────────────────────────────────────────────────────
# Status
# ──────────────────────────────────────────────────────────────────

def get_data_status(session: Session, major_code: str) -> dict:
    pr_dv = get_active_dataset(session, major_code, 'progress_report')
    cc_dv = get_active_dataset(session, major_code, 'course_config')

    pr_status: dict = {'has_report': pr_dv is not None, 'student_count': 0, 'uploaded_at': None}
    if pr_dv and pr_dv.parsed_payload:
        records = pr_dv.parsed_payload.get('records', [])
        ids = {r.get('ID') for r in records if r.get('ID')}
        pr_status['student_count'] = len(ids)
        if pr_dv.created_at:
            pr_status['uploaded_at'] = pr_dv.created_at.isoformat()

    cc_status: dict = {'has_config': cc_dv is not None, 'required_count': 0, 'intensive_count': 0}
    if cc_dv and cc_dv.parsed_payload:
        config = cc_dv.parsed_payload.get('records', [{}])[0]
        cc_status['required_count'] = len(config.get('target_courses', {}))
        cc_status['intensive_count'] = len(config.get('intensive_courses', {}))

    return {'progress_report': pr_status, 'course_config': cc_status}


# ──────────────────────────────────────────────────────────────────
# Report generation
# ──────────────────────────────────────────────────────────────────

def generate_report(
    session: Session,
    major_code: str,
    show_all_grades: bool = False,
    page: int = 1,
    page_size: int = 50,
    search: str = '',
) -> ReportResponse:
    config = _load_course_config(session, major_code)
    if config is None:
        raise LookupError("No course configuration uploaded yet.")

    df = _load_progress_df(session, major_code)
    if df is None:
        raise LookupError("No progress report uploaded yet.")

    equiv_map = _build_equiv_map(session, major_code)
    assign_map = _build_assign_map(session, major_code)

    req_df, int_df, _extra_df, extra_list = process_progress_report(
        df,
        target_courses=config.get('target_courses', {}),
        intensive_courses=config.get('intensive_courses', {}),
        target_rules=config.get('target_rules', {}),
        intensive_rules=config.get('intensive_rules', {}),
        per_student_assignments=assign_map or None,
        equivalent_courses_mapping=equiv_map or None,
    )

    target_courses: dict[str, int] = config.get('target_courses', {})
    intensive_courses: dict[str, int] = config.get('intensive_courses', {})
    total_students = len(req_df)

    # Search filter
    if search:
        s = search.strip().lower()
        mask = (
            req_df['ID'].str.lower().str.contains(s, na=False)
            | req_df['NAME'].str.lower().str.contains(s, na=False)
        )
        req_df = req_df[mask].reset_index(drop=True)
        int_df = int_df[int_df['ID'].isin(req_df['ID'])].reset_index(drop=True)
        total_students = len(req_df)

    # Pagination
    start = (page - 1) * page_size
    req_page = req_df.iloc[start: start + page_size]
    page_ids = set(req_page['ID'].tolist())
    int_page = int_df[int_df['ID'].isin(page_ids)]

    required_rows = _build_student_rows(req_page, target_courses, show_all_grades)
    intensive_rows = _build_student_rows(int_page, intensive_courses, show_all_grades)

    return ReportResponse(
        required=required_rows,
        intensive=intensive_rows,
        extra_courses=extra_list,
        total_students=total_students,
        page=page,
        page_size=page_size,
    )


def _build_student_rows(
    df: pd.DataFrame,
    courses_dict: dict[str, int],
    show_all_grades: bool,
) -> list[StudentProgressRow]:
    if df.empty:
        return []

    course_cols = list(courses_dict.keys())
    gpa_map = calculate_gpa_for_rows(df, courses_dict)

    rows: list[StudentProgressRow] = []
    for _, row in df.iterrows():
        courses: dict[str, str] = {}
        for course in course_cols:
            val = str(row.get(course, 'NR'))
            if show_all_grades:
                courses[course] = val
            else:
                courses[course] = extract_primary_grade(val) if val not in ('NR', '') else 'NR'

        credits = calculate_credits(row, courses_dict)
        rows.append(
            StudentProgressRow(
                student_id=str(row['ID']),
                name=str(row['NAME']),
                courses=courses,
                completed_credits=credits['completed'],
                registered_credits=credits['registered'],
                remaining_credits=credits['remaining'],
                total_credits=credits['total'],
                gpa=gpa_map.get(str(row['ID'])),
            )
        )
    return rows


# ──────────────────────────────────────────────────────────────────
# Excel export helpers
# ──────────────────────────────────────────────────────────────────

def _collapse_pass_fail(val: str) -> str:
    """Mirror the frontend collapsePassFail logic: returns 'c', 'cr', or 'nc'."""
    if not val or val == 'NR':
        return 'nc'
    if val.upper().startswith('CR'):
        return 'cr'
    for entry in val.split(','):
        entry = entry.strip()
        if entry.upper().startswith('CR'):
            return 'cr'
    for entry in val.split(','):
        parts = entry.strip().split('|')
        if len(parts) == 2:
            right = parts[1].strip().upper()
            try:
                n = int(right)
                if n > 0:
                    return 'c'
                return 'nc'
            except ValueError:
                if right == 'PASS':
                    return 'c'
    return 'nc'


# ──────────────────────────────────────────────────────────────────
# Excel export
# ──────────────────────────────────────────────────────────────────

def export_report_excel(
    session: Session,
    major_code: str,
    show_all_grades: bool = False,
    collapse_mode: bool = False,
    advising_format: bool = False,
) -> bytes:
    config = _load_course_config(session, major_code)
    if config is None:
        raise LookupError("No course configuration uploaded yet.")

    df = _load_progress_df(session, major_code)
    if df is None:
        raise LookupError("No progress report uploaded yet.")

    equiv_map = _build_equiv_map(session, major_code)
    assign_map = _build_assign_map(session, major_code)

    req_df, int_df, _extra_df, _extra_list = process_progress_report(
        df,
        target_courses=config.get('target_courses', {}),
        intensive_courses=config.get('intensive_courses', {}),
        target_rules=config.get('target_rules', {}),
        intensive_rules=config.get('intensive_rules', {}),
        per_student_assignments=assign_map or None,
        equivalent_courses_mapping=equiv_map or None,
    )

    wb = openpyxl.Workbook()
    _write_excel_sheet(
        wb.active, 'Required', req_df,
        config.get('target_courses', {}), show_all_grades, collapse_mode, advising_format,
    )
    ws2 = wb.create_sheet('Intensive')
    _write_excel_sheet(ws2, 'Intensive', int_df, config.get('intensive_courses', {}), show_all_grades, collapse_mode, advising_format)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_excel_sheet(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    title: str,
    df: pd.DataFrame,
    courses_dict: dict[str, int],
    show_all_grades: bool,
    collapse_mode: bool = False,
    advising_format: bool = False,
) -> None:
    ws.title = title
    course_cols = list(courses_dict.keys())
    # Use column names matching the advising app's expected format when pushing
    if advising_format:
        summary_headers = ['# of Credits Completed', '# Registered', '# Remaining', 'Total Credits', 'GPA']
    else:
        summary_headers = ['Completed', 'Registered', 'Remaining', 'Total', 'GPA']
    headers = ['ID', 'NAME'] + course_cols + summary_headers
    ws.append(headers)

    if df.empty:
        return

    gpa_map = calculate_gpa_for_rows(df, courses_dict)

    for _, row in df.iterrows():
        display: dict[str, str] = {}
        for course in course_cols:
            val = str(row.get(course, 'NR'))
            if collapse_mode:
                primary = extract_primary_grade(val) if val not in ('NR', '') else 'NR'
                display[course] = _collapse_pass_fail(primary)
            else:
                display[course] = (
                    val if show_all_grades
                    else (extract_primary_grade(val) if val not in ('NR', '') else 'NR')
                )

        credits = calculate_credits(row, courses_dict)
        gpa = gpa_map.get(str(row['ID']))
        ws_row = [str(row['ID']), str(row['NAME'])]
        ws_row += [display[c] for c in course_cols]
        ws_row += [
            credits['completed'], credits['registered'],
            credits['remaining'], credits['total'],
            round(float(gpa), 2) if gpa is not None else '',
        ]
        ws.append(ws_row)

    # Apply background colours
    for r_idx, xl_row in enumerate(ws.iter_rows(min_row=2), start=2):
        for c_idx, cell in enumerate(xl_row):
            if c_idx < 2:
                continue
            col_header = headers[c_idx] if c_idx < len(headers) else ''
            if col_header not in courses_dict:
                continue
            raw_css = cell_color(str(cell.value) if cell.value else 'NR')
            # raw_css is like "background-color: #28a745"
            hex_col = raw_css.replace('background-color:', '').strip().lstrip('#')
            if hex_col:
                cell.fill = PatternFill(
                    start_color=hex_col.upper(),
                    end_color=hex_col.upper(),
                    fill_type='solid',
                )


# ──────────────────────────────────────────────────────────────────
# Push-to-Advising
# ──────────────────────────────────────────────────────────────────

def push_progress_to_advising(
    session: Session,
    major_code: str,
    user_id: Optional[int],
) -> dict[str, object]:
    """Generate a collapsed c/cr/nc Excel and store it as the 'progress' dataset
    in the advising app for the given major."""
    xlsx_bytes = export_report_excel(
        session, major_code,
        show_all_grades=False,
        collapse_mode=True,
        advising_format=True,
    )
    version = upload_dataset(
        session,
        major_code=major_code,
        dataset_type='progress',
        filename=f'progress_{major_code}.xlsx',
        content=xlsx_bytes,
        user_id=user_id,
    )

    # Link the new progress dataset to the currently active advising period
    # (Note: upload_dataset already calls _link_to_active_period for 'progress',
    #  but we also explicitly set it here for clarity and backward compat)
    major = session.scalar(select(Major).where(Major.code == major_code))
    if major:
        active_period = session.scalar(
            select(AdvisingPeriod).where(
                AdvisingPeriod.major_id == major.id,
                AdvisingPeriod.is_active.is_(True),
            )
        )
        if active_period:
            active_period.progress_dataset_version_id = version.id
            session.commit()

    student_count = len(version.parsed_payload.get('records', []))
    return {'message': f'Progress report pushed for {major_code} ({student_count} students).', 'version_id': version.id, 'student_count': student_count}


# ──────────────────────────────────────────────────────────────────
# Bulk Assignment Upload
# ──────────────────────────────────────────────────────────────────

def bulk_upsert_assignments_from_excel(
    session: Session,
    major_code: str,
    content: bytes,
) -> dict[str, object]:
    """Parse an Excel file with columns Student ID, Assignment Type, Course Code
    and upsert each row. Returns {upserted, skipped, errors}."""
    try:
        df = pd.read_excel(io.BytesIO(content))
    except Exception as exc:
        raise ValueError(f'Could not parse Excel file: {exc}') from exc

    # Normalise column names: strip whitespace, lower, replace spaces with underscore
    df.columns = [str(c).strip().lower().replace(' ', '_') for c in df.columns]

    # Accept both 'id' and 'student_id' for the student ID column
    if 'student_id' not in df.columns and 'id' in df.columns:
        df = df.rename(columns={'id': 'student_id'})

    required = {'student_id', 'assignment_type', 'course_code'}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(
            f'Missing required columns: {", ".join(sorted(missing))}. '
            f'Expected: Student ID (or ID), Assignment Type, Course Code.'
        )

    upserted = 0
    skipped = 0
    errors: list[str] = []

    for i, row in df.iterrows():
        student_id = str(row['student_id']).strip()
        assignment_type = str(row['assignment_type']).strip()
        course_code = str(row['course_code']).strip()

        if not student_id or not assignment_type or not course_code:
            errors.append(f'Row {i + 2}: empty value(s) — skipped.')
            skipped += 1
            continue

        try:
            upsert_assignment(session, major_code, student_id, assignment_type, course_code)
            session.flush()
            upserted += 1
        except (LookupError, ValueError) as exc:
            errors.append(f'Row {i + 2} ({student_id}): {exc}')
            skipped += 1

    return {'upserted': upserted, 'skipped': skipped, 'errors': errors}


# ──────────────────────────────────────────────────────────────────
# Student exemptions (e.g. ARAB201)
# ──────────────────────────────────────────────────────────────────

def list_exemptions(
    session: Session, major_code: str, student_id: Optional[str] = None
) -> list:
    from app.models import StudentExemption
    major_id = _get_major_id(session, major_code)
    q = select(StudentExemption).where(StudentExemption.major_id == major_id)
    if student_id:
        q = q.where(StudentExemption.student_id == str(student_id))
    return list(session.scalars(q).all())


def set_exemption(
    session: Session, major_code: str, student_id: str, course_code: str
):
    """Grant an exemption. Idempotent (no-op if it already exists)."""
    from app.models import StudentExemption
    major_id = _get_major_id(session, major_code)
    existing = session.scalar(
        select(StudentExemption).where(
            StudentExemption.major_id == major_id,
            StudentExemption.student_id == str(student_id),
            StudentExemption.course_code == str(course_code),
        )
    )
    if existing:
        return existing
    exemption = StudentExemption(
        major_id=major_id,
        student_id=str(student_id),
        course_code=str(course_code),
    )
    session.add(exemption)
    session.flush()
    return exemption


def remove_exemption(
    session: Session, major_code: str, student_id: str, course_code: str
) -> None:
    """Remove an exemption. Silent no-op if it doesn't exist."""
    from app.models import StudentExemption
    major_id = _get_major_id(session, major_code)
    row = session.scalar(
        select(StudentExemption).where(
            StudentExemption.major_id == major_id,
            StudentExemption.student_id == str(student_id),
            StudentExemption.course_code == str(course_code),
        )
    )
    if row:
        session.delete(row)
        session.flush()
