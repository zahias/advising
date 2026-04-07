from __future__ import annotations

from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any, Iterable

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models import AuditEvent, Bypass, Major, SessionSnapshot, StudentSelection
from app.schemas.insights import CourseOfferingRecommendation, DashboardMetrics
from app.services.audit import log_event
from app.services.dataset_service import dataset_dataframe
from app.services.period_service import current_period

def _find_legacy_root() -> Path:
    """Walk up from this file until we find eligibility_utils.py (the workspace root)."""
    d = Path(__file__).resolve().parent
    for _ in range(8):
        if (d / 'eligibility_utils.py').exists():
            return d
        d = d.parent
    return Path(__file__).resolve().parents[4]  # fallback

ROOT_DIR = _find_legacy_root()
import sys
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

from eligibility_utils import (  # noqa: E402
    build_requisites_str,
    check_course_completed,
    check_course_registered,
    check_eligibility,
    get_corequisite_and_concurrent_courses,
    get_mutual_concurrent_pairs,
    get_student_standing,
    parse_requirements,
    parse_requirements_grouped,
)
from reporting import add_summary_sheet, apply_full_report_formatting, apply_individual_compact_formatting  # noqa: E402

STATUS_CODES = {
    'completed': 'c',
    'registered': 'r',
    'advised': 'a',
    'advised_repeat': 'ar',
    'optional': 'o',
    'bypass': 'b',
    'eligible': 'na',
    'not_eligible': 'ne',
}

DEFAULT_SEMESTER_FILTER = 'All Courses'

HEADER_FILL = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF')
TITLE_FONT = Font(bold=True, size=14)
SUBTITLE_FONT = Font(italic=True, color='4B5563')

ACTION_COLORS = {
    'completed': 'C6E0B4',
    'registered': 'BDD7EE',
    'advised': 'FFF2CC',
    'eligible (bypass)': 'E9D5FF',
    'eligible not chosen': 'E1F0FF',
    'not eligible': 'F8CECC',
}


def _style_header_row(ws, row_index: int) -> None:
    for cell in ws[row_index]:
        if cell.value is None:
            continue
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')


def _autosize_columns(ws, *, max_width: int = 72) -> None:
    for column_cells in ws.columns:
        max_length = 0
        for cell in column_cells:
            if cell.value is None:
                continue
            max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max_length + 2, max_width)


def _apply_action_column_colors(ws, *, header_row: int) -> None:
    action_col = None
    for idx, cell in enumerate(ws[header_row], start=1):
        if str(cell.value or '').strip().lower() == 'action':
            action_col = idx
            break
    if action_col is None:
        return

    for row in range(header_row + 1, ws.max_row + 1):
        action_value = str(ws.cell(row=row, column=action_col).value or '').strip().lower()
        color = ACTION_COLORS.get(action_value)
        if color:
            ws.cell(row=row, column=action_col).fill = PatternFill(start_color=color, end_color=color, fill_type='solid')


def _major_or_error(session: Session, major_code: str) -> Major:
    major = session.scalar(select(Major).where(Major.code == major_code))
    if not major:
        raise ValueError(f'Unknown major: {major_code}')
    return major


def _courses_df(session: Session, major_code: str) -> pd.DataFrame:
    return dataset_dataframe(session, major_code, 'courses')


def _progress_df(session: Session, major_code: str) -> pd.DataFrame:
    return dataset_dataframe(session, major_code, 'progress')


def _selection_map(session: Session, major: Major, period_id: int | None) -> dict[str, StudentSelection]:
    if not period_id:
        return {}
    return {
        item.student_id: item
        for item in session.scalars(
            select(StudentSelection).where(StudentSelection.major_id == major.id, StudentSelection.period_id == period_id)
        ).all()
    }


def _bypass_map(session: Session, major: Major) -> dict[tuple[str, str], dict[str, Any]]:
    bypasses = session.scalars(select(Bypass).where(Bypass.major_id == major.id)).all()
    return {(item.student_id, item.course_code): {'note': item.note, 'advisor': item.advisor_name} for item in bypasses}


def _semester_filter_options(courses_df: pd.DataFrame) -> list[str]:
    if 'Suggested Semester' not in courses_df.columns:
        return [DEFAULT_SEMESTER_FILTER]
    values = []
    semester_types = set()
    for value in courses_df['Suggested Semester'].dropna().tolist():
        label = str(value).strip()
        if not label or label.lower() in {'nan', 'none'}:
            continue
        values.append(label)
        parts = label.split('-')
        if len(parts) == 2:
            semester_types.add(parts[0])
    return [DEFAULT_SEMESTER_FILTER, *sorted(semester_types), *sorted(set(values))]


def _filter_courses_by_semester(courses_df: pd.DataFrame, course_list: list[str], semester_filter: str) -> list[str]:
    if semester_filter == DEFAULT_SEMESTER_FILTER or 'Suggested Semester' not in courses_df.columns:
        return course_list
    filtered: list[str] = []
    for course_code in course_list:
        course_info = courses_df.loc[courses_df['Course Code'].astype(str) == str(course_code)]
        if course_info.empty:
            continue
        suggested_semester = str(course_info.iloc[0].get('Suggested Semester', '')).strip()
        if not suggested_semester or suggested_semester.lower() in {'nan', 'none'}:
            continue
        if suggested_semester == semester_filter:
            filtered.append(course_code)
        elif '-' not in semester_filter and suggested_semester.startswith(f'{semester_filter}-'):
            filtered.append(course_code)
    return filtered


def _course_code_lists(courses_df: pd.DataFrame, semester_filter: str = DEFAULT_SEMESTER_FILTER) -> tuple[list[str], list[str]]:
    type_series = courses_df.get('Type', pd.Series(dtype=str)).astype(str).str.strip().str.lower()
    required_courses = courses_df.loc[type_series == 'required', 'Course Code'].dropna().astype(str).tolist()
    intensive_courses = courses_df.loc[type_series == 'intensive', 'Course Code'].dropna().astype(str).tolist()
    if semester_filter != DEFAULT_SEMESTER_FILTER:
        required_courses = _filter_courses_by_semester(courses_df, required_courses, semester_filter)
        intensive_courses = _filter_courses_by_semester(courses_df, intensive_courses, semester_filter)
    return required_courses, intensive_courses


def _course_metadata(courses_df: pd.DataFrame) -> dict[str, dict[str, Any]]:
    metadata: dict[str, dict[str, Any]] = {}
    for _, row in courses_df.iterrows():
        course_code = str(row.get('Course Code', '')).strip()
        if not course_code:
            continue
        metadata[course_code] = {
            'course_code': course_code,
            'title': str(row.get('Course Title', row.get('Title', '')) or course_code),
            'course_type': str(row.get('Type', '') or ''),
            'credits': float(row.get('Credits', 0) or 0),
            'requisites': build_requisites_str(row),
            'suggested_semester': str(row.get('Suggested Semester', '') or '').strip(),
        }
    return metadata


def _status_code(
    student_row: pd.Series,
    student_id: str,
    course_code: str,
    courses_df: pd.DataFrame,
    advised: list[str],
    optional: list[str],
    repeat: list[str],
    mutual_pairs: dict[str, list[str]],
    bypasses: dict[tuple[str, str], dict[str, Any]],
    simulated_for_student: list[str] | None = None,
) -> str:
    simulated_for_student = simulated_for_student or []
    if course_code in repeat:
        return 'ar'
    if check_course_completed(student_row, course_code):
        return 'c'
    if check_course_registered(student_row, course_code):
        return 'r'
    if course_code in simulated_for_student:
        return 's'
    if course_code in optional:
        return 'o'
    if course_code in advised:
        return 'a'
    status, _ = check_eligibility(
        student_row,
        course_code,
        advised,
        courses_df,
        registered_courses=simulated_for_student,
        ignore_offered=True,
        mutual_pairs=mutual_pairs,
        bypass_map={course_code: bypasses.get((student_id, course_code), {})} if (student_id, course_code) in bypasses else {},
    )
    if status == 'Eligible (Bypass)':
        return 'b'
    return 'na' if status == 'Eligible' else 'ne'


def _simulate_registrations(
    progress_df: pd.DataFrame,
    courses_df: pd.DataFrame,
    selections_by_student: dict[str, StudentSelection],
    simulated_courses: list[str],
    mutual_pairs: dict[str, list[str]],
    bypasses: dict[tuple[str, str], dict[str, Any]],
) -> dict[str, list[str]]:
    if not simulated_courses:
        return {}
    original_rows = {
        str(row.get('ID')): row
        for _, row in progress_df.iterrows()
    }
    simulated_map: dict[str, list[str]] = {}
    for student_id, student_row in original_rows.items():
        selection = selections_by_student.get(student_id)
        advised = sorted(set((selection.advised if selection else []) + (selection.optional if selection else [])))
        simulated_map[student_id] = []
        max_iterations = len(simulated_courses)
        for _ in range(max_iterations):
            added_this_iteration = False
            for sim_course in simulated_courses:
                if sim_course in simulated_map[student_id]:
                    continue
                status, _ = check_eligibility(
                    student_row,
                    sim_course,
                    advised,
                    courses_df,
                    registered_courses=simulated_map[student_id],
                    ignore_offered=True,
                    mutual_pairs=mutual_pairs,
                    bypass_map={sim_course: bypasses.get((student_id, sim_course), {})} if (student_id, sim_course) in bypasses else {},
                )
                if status in {'Eligible', 'Eligible (Bypass)'}:
                    simulated_map[student_id].append(sim_course)
                    added_this_iteration = True
            if not added_this_iteration:
                break
    return simulated_map


def _planner_entity_id(major_code: str, period_code: str | None) -> str:
    return f'{major_code}:{period_code or "no-period"}'


def dashboard_metrics(session: Session, major_code: str) -> DashboardMetrics:
    major = _major_or_error(session, major_code)
    progress_df = _progress_df(session, major_code)
    if progress_df.empty:
        return DashboardMetrics(total_students=0, advised_students=0, not_advised_students=0, progress_percent=0, graduating_soon_unadvised=[], recent_activity=[])
    period = current_period(session, major_code)
    advised_ids = set()
    recent_activity: list[dict[str, Any]] = []
    if period:
        latest_sessions = list(session.scalars(select(SessionSnapshot).where(SessionSnapshot.major_id == major.id, SessionSnapshot.period_id == period.id).order_by(SessionSnapshot.created_at.desc())))
        for snap in latest_sessions:
            advised_ids.add(str(snap.student_id))
        for snap in latest_sessions[:5]:
            recent_activity.append({'student_name': snap.payload.get('student_name', snap.student_id), 'created_at': snap.created_at.isoformat()})
    total_students = len(progress_df)
    advised_count = 0
    graduating_soon_unadvised: list[dict[str, str]] = []
    for _, row in progress_df.iterrows():
        student_id = str(row.get('ID'))
        remaining = float(row.get('# Remaining', row.get('Remaining Credits', 999)) or 999)
        if student_id in advised_ids:
            advised_count += 1
        elif remaining <= 36:
            graduating_soon_unadvised.append(
                {
                    'student_id': student_id,
                    'student_name': str(row.get('NAME', student_id)),
                }
            )
    percent = int((advised_count / total_students) * 100) if total_students else 0
    try:
        rem_col = '# Remaining' if '# Remaining' in progress_df.columns else 'Remaining Credits'
        if rem_col in progress_df.columns and 'ID' in progress_df.columns:
            rem_series = pd.to_numeric(progress_df.groupby('ID')[rem_col].first(), errors='coerce').dropna()
            credit_distribution: list[dict[str, Any]] = [
                {'label': '≤18', 'count': int((rem_series <= 18).sum())},
                {'label': '19–36', 'count': int(((rem_series > 18) & (rem_series <= 36)).sum())},
                {'label': '37–72', 'count': int(((rem_series > 36) & (rem_series <= 72)).sum())},
                {'label': '73+', 'count': int((rem_series > 72).sum())},
            ]
        else:
            credit_distribution = []
    except Exception:
        credit_distribution = []
    return DashboardMetrics(
        total_students=total_students,
        advised_students=advised_count,
        not_advised_students=total_students - advised_count,
        progress_percent=percent,
        graduating_soon_unadvised=graduating_soon_unadvised[:5],
        recent_activity=recent_activity,
        credit_distribution=credit_distribution,
    )


def _build_prerequisite_map(courses_df: pd.DataFrame) -> dict[str, list[str]]:
    prereq_map: dict[str, list[str]] = {}
    for _, row in courses_df.iterrows():
        course_code = str(row.get('Course Code', ''))
        # Use grouped parser so that OR-alternatives (separated by ' / ') each
        # map to this course individually — either one can unlock it.
        for col in ('Prerequisite', 'Concurrent', 'Corequisite'):
            for or_group in parse_requirements_grouped(row.get(col)):
                for alt in or_group:
                    prereq_map.setdefault(alt, []).append(course_code)
    return prereq_map


def _calculate_cascading_eligibility(courses_df: pd.DataFrame, progress_df: pd.DataFrame, offered_course: str, mutual_pairs: dict[str, list[str]]) -> int:
    unlocked_students = 0
    for _, student in progress_df.iterrows():
        if check_course_completed(student, offered_course) or check_course_registered(student, offered_course):
            continue
        simulated = [offered_course]
        for _, course_row in courses_df.iterrows():
            target_course = str(course_row.get('Course Code', ''))
            if target_course == offered_course:
                continue
            status, _ = check_eligibility(student, target_course, [], courses_df, registered_courses=simulated, ignore_offered=True, mutual_pairs=mutual_pairs)
            if status == 'Eligible':
                unlocked_students += 1
                break
    return unlocked_students


def course_offering_recommendations(session: Session, major_code: str, graduation_threshold: int = 30, min_eligible_students: int = 3) -> list[CourseOfferingRecommendation]:
    courses_df = _courses_df(session, major_code)
    progress_df = _progress_df(session, major_code)
    if courses_df.empty or progress_df.empty:
        return []
    progress_df = progress_df.copy()
    progress_df['Remaining Credits'] = pd.to_numeric(progress_df.get('# Remaining', 0), errors='coerce').fillna(0)
    mutual_pairs = get_mutual_concurrent_pairs(courses_df)
    prereq_map = _build_prerequisite_map(courses_df)
    recommendations: list[CourseOfferingRecommendation] = []
    for _, course_row in courses_df.iterrows():
        course = str(course_row.get('Course Code', ''))
        if not course:
            continue
        eligible = 0
        graduating = 0
        for _, student in progress_df.iterrows():
            if check_course_completed(student, course) or check_course_registered(student, course):
                continue
            status, _ = check_eligibility(student, course, [], courses_df, registered_courses=[], ignore_offered=True, mutual_pairs=mutual_pairs)
            if status == 'Eligible':
                eligible += 1
                if float(student.get('Remaining Credits', 999) or 999) <= graduation_threshold:
                    graduating += 1
        if eligible < min_eligible_students:
            continue
        bottleneck = len(prereq_map.get(course, []))
        cascading = _calculate_cascading_eligibility(courses_df, progress_df, course, mutual_pairs)
        score = eligible + (graduating * 3) + (bottleneck * 2) + (cascading * 1.5)
        reason_parts: list[str] = []
        if graduating:
            reason_parts.append(f'{graduating} graduating')
        if bottleneck:
            reason_parts.append(f'unlocks {bottleneck} courses')
        if cascading:
            reason_parts.append(f'+{cascading} cascading eligible')
        recommendations.append(
            CourseOfferingRecommendation(
                course=course,
                priority_score=float(score),
                currently_eligible=eligible,
                graduating_students=graduating,
                bottleneck_score=bottleneck,
                cascading_eligible=cascading,
                reason='; '.join(reason_parts) if reason_parts else 'General progression',
            )
        )
    return sorted(recommendations, key=lambda item: item.priority_score, reverse=True)


def saved_course_offering_plan(session: Session, major_code: str) -> dict[str, Any]:
    period = current_period(session, major_code)
    entity_id = _planner_entity_id(major_code, period.period_code if period else None)
    event = session.scalar(
        select(AuditEvent)
        .where(
            AuditEvent.entity_type == 'course_offering_plan',
            AuditEvent.entity_id == entity_id,
        )
        .order_by(AuditEvent.created_at.desc())
    )
    if not event:
        return {
            'selected_courses': [],
            'graduation_threshold': 30,
            'min_eligible_students': 3,
            'total_eligible': 0,
            'total_graduating': 0,
            'saved_at': None,
        }
    payload = dict(event.payload or {})
    payload['saved_at'] = event.created_at.isoformat() if event.created_at else None
    payload.setdefault('selected_courses', [])
    payload.setdefault('graduation_threshold', 30)
    payload.setdefault('min_eligible_students', 3)
    payload.setdefault('total_eligible', 0)
    payload.setdefault('total_graduating', 0)
    return payload


def save_course_offering_plan(
    session: Session,
    major_code: str,
    *,
    selected_courses: list[str],
    graduation_threshold: int,
    min_eligible_students: int,
    actor_user_id: int | None,
) -> dict[str, Any]:
    recommendations = course_offering_recommendations(
        session,
        major_code,
        graduation_threshold=graduation_threshold,
        min_eligible_students=min_eligible_students,
    )
    selected_set = set(selected_courses)
    valid_selected = [item.course for item in recommendations if item.course in selected_set]
    total_eligible = sum(item.currently_eligible for item in recommendations if item.course in selected_set)
    total_graduating = sum(item.graduating_students for item in recommendations if item.course in selected_set)
    period = current_period(session, major_code)
    payload = {
        'selected_courses': valid_selected,
        'graduation_threshold': int(graduation_threshold),
        'min_eligible_students': int(min_eligible_students),
        'total_eligible': int(total_eligible),
        'total_graduating': int(total_graduating),
    }
    log_event(
        session,
        actor_user_id=actor_user_id,
        event_type='planner_selection_saved',
        entity_type='course_offering_plan',
        entity_id=_planner_entity_id(major_code, period.period_code if period else None),
        payload=payload,
    )
    session.commit()
    payload['saved_at'] = datetime.utcnow().isoformat()
    return payload


def all_students_view(
    session: Session,
    major_code: str,
    *,
    simulated_courses: list[str] | None = None,
    semester_filter: str = DEFAULT_SEMESTER_FILTER,
) -> dict[str, Any]:
    courses_df = _courses_df(session, major_code)
    progress_df = _progress_df(session, major_code)
    major = _major_or_error(session, major_code)
    if courses_df.empty or progress_df.empty:
        return {
            'rows': [],
            'required_courses': [],
            'intensive_courses': [],
            'course_metadata': {},
            'simulation_options': [],
            'semester_options': [DEFAULT_SEMESTER_FILTER],
            'legend': [],
            'remaining_range': {'min': 0, 'max': 0},
            'simulated_courses': [],
        }
    period = current_period(session, major_code)
    selections_by_student = _selection_map(session, major, period.id if period else None)
    bypass_map = _bypass_map(session, major)
    mutual_pairs = get_mutual_concurrent_pairs(courses_df)
    simulation_options = get_corequisite_and_concurrent_courses(courses_df)
    semester_options = _semester_filter_options(courses_df)
    required_courses, intensive_courses = _course_code_lists(courses_df, semester_filter)
    simulated_courses = [course for course in (simulated_courses or []) if course in simulation_options]
    simulated_map = _simulate_registrations(progress_df, courses_df, selections_by_student, simulated_courses, mutual_pairs, bypass_map)
    metadata = _course_metadata(courses_df)
    all_courses = [str(code) for code in courses_df['Course Code'].dropna().tolist()]
    rows: list[dict[str, Any]] = []
    working = progress_df.copy()
    working['Remaining Credits'] = pd.to_numeric(working.get('# Remaining', working.get('Remaining Credits', 0)), errors='coerce').fillna(0).astype(int)
    min_remaining = int(working['Remaining Credits'].min()) if not working.empty else 0
    max_remaining = int(working['Remaining Credits'].max()) if not working.empty else 0
    for _, student in working.iterrows():
        sid = str(student.get('ID'))
        total = float(student.get('# of Credits Completed', 0) or 0) + float(student.get('# Registered', 0) or 0)
        selection = selections_by_student.get(sid)
        advised = list(selection.advised if selection else [])
        optional = list(selection.optional if selection else [])
        repeat = list(selection.repeat if selection else [])
        note = selection.note if selection else ''
        row = {
            'student_id': sid,
            'student_name': str(student.get('NAME', sid)),
            'standing': get_student_standing(total),
            'total_credits': total,
            'remaining_credits': float(student.get('Remaining Credits', student.get('# Remaining', 0)) or 0),
            'advising_status': 'Advised' if (advised or optional or repeat or note) else 'Not Advised',
            'courses': {},
        }
        for course in all_courses:
            row['courses'][course] = _status_code(
                student,
                sid,
                course,
                courses_df,
                advised=advised,
                optional=optional,
                repeat=repeat,
                mutual_pairs=mutual_pairs,
                bypasses=bypass_map,
                simulated_for_student=simulated_map.get(sid, []),
            )
        rows.append(row)
    return {
        'rows': rows,
        'required_courses': required_courses,
        'intensive_courses': intensive_courses,
        'course_metadata': metadata,
        'simulation_options': simulation_options,
        'semester_options': semester_options,
        'legend': [
            {'code': 'c', 'label': 'Completed'},
            {'code': 'r', 'label': 'Registered'},
            {'code': 's', 'label': 'Simulated'},
            {'code': 'a', 'label': 'Advised'},
            {'code': 'ar', 'label': 'Advised Repeat'},
            {'code': 'o', 'label': 'Optional'},
            {'code': 'b', 'label': 'Bypass'},
            {'code': 'na', 'label': 'Eligible not chosen'},
            {'code': 'ne', 'label': 'Not eligible'},
        ],
        'remaining_range': {'min': min_remaining, 'max': max_remaining},
        'simulated_courses': simulated_courses,
    }


def individual_student_view(session: Session, major_code: str, student_id: str, selected_courses: list[str] | None = None) -> dict[str, Any]:
    courses_df = _courses_df(session, major_code)
    progress_df = _progress_df(session, major_code)
    major = _major_or_error(session, major_code)
    if courses_df.empty or progress_df.empty:
        raise ValueError('Courses and progress datasets must be uploaded first')
    student_match = progress_df.loc[progress_df['ID'].astype(str) == str(student_id)]
    if student_match.empty:
        raise ValueError(f'Student not found: {student_id}')
    student = student_match.iloc[0]
    period = current_period(session, major_code)
    selections_by_student = _selection_map(session, major, period.id if period else None)
    selection = selections_by_student.get(str(student_id))
    advised = list(selection.advised if selection else [])
    optional = list(selection.optional if selection else [])
    repeat = list(selection.repeat if selection else [])
    courses = selected_courses or [str(code) for code in courses_df['Course Code'].dropna().tolist()]
    mutual_pairs = get_mutual_concurrent_pairs(courses_df)
    bypass_map = _bypass_map(session, major)
    rows: list[dict[str, Any]] = []
    status_map: dict[str, str] = {}
    for course in courses:
        if course in repeat:
            code = 'ar'
        elif check_course_completed(student, course):
            code = 'c'
        elif check_course_registered(student, course):
            code = 'r'
        elif course in optional:
            code = 'o'
        elif course in advised:
            code = 'a'
        else:
            status, justification = check_eligibility(
                student,
                course,
                advised,
                courses_df,
                registered_courses=[],
                ignore_offered=True,
                mutual_pairs=mutual_pairs,
                bypass_map={course: bypass_map.get((str(student_id), course), {})} if (str(student_id), course) in bypass_map else {},
            )
            code = 'b' if status == 'Eligible (Bypass)' else 'na' if status == 'Eligible' else 'ne'
        status_map[course] = code
        rows.append({'course_code': course, 'status_code': code})
    return {
        'student_id': str(student_id),
        'student_name': str(student.get('NAME', student_id)),
        'selected_courses': courses,
        'statuses': status_map,
        'advised': advised,
        'optional': optional,
        'repeat': repeat,
        'note': selection.note if selection else '',
    }


def qaa_sheet(session: Session, major_code: str, graduating_threshold: int = 36) -> list[dict[str, Any]]:
    courses_df = _courses_df(session, major_code)
    progress_df = _progress_df(session, major_code).copy()
    major = _major_or_error(session, major_code)
    if courses_df.empty or progress_df.empty:
        return []
    progress_df['ID'] = pd.to_numeric(progress_df['ID'], errors='coerce')
    progress_df = progress_df.dropna(subset=['ID'])
    progress_df['ID'] = progress_df['ID'].astype(int)
    progress_df['Remaining Credits'] = pd.to_numeric(progress_df.get('# Remaining', 0), errors='coerce').fillna(0).astype(int)
    period = current_period(session, major_code)
    selections_by_student = _selection_map(session, major, period.id if period else None)
    mutual_pairs = get_mutual_concurrent_pairs(courses_df)
    all_bypasses = _bypass_map(session, major)
    qaa_data: list[dict[str, Any]] = []
    for course_code in courses_df['Course Code'].dropna().unique().tolist():
        course_info = courses_df.loc[courses_df['Course Code'] == course_code]
        if course_info.empty:
            continue
        course_name = str(course_info.iloc[0].get('Course Name', course_info.iloc[0].get('Title', '')))
        eligible_students: list[int] = []
        advised_students: list[int] = []
        optional_advised_students: list[int] = []
        not_advised_students: list[int] = []
        skipped_advising_students: list[int] = []
        attended_graduating: list[int] = []
        skipped_graduating: list[int] = []
        for _, student in progress_df.iterrows():
            sid = int(student['ID'])
            remaining = int(student.get('Remaining Credits', 999) or 999)
            is_graduating = remaining <= graduating_threshold
            if check_course_completed(student, course_code) or check_course_registered(student, course_code):
                continue
            sel = selections_by_student.get(str(sid))
            advised = sel.advised if sel else []
            optional = sel.optional if sel else []
            repeat = sel.repeat if sel else []
            note = sel.note if sel else ''
            student_bypasses = {
                cc: info for (student_key, cc), info in all_bypasses.items() if student_key == str(sid)
            }
            status, _ = check_eligibility(
                student,
                str(course_code),
                advised,
                courses_df,
                registered_courses=[],
                ignore_offered=True,
                mutual_pairs=mutual_pairs,
                bypass_map=student_bypasses,
            )
            if status not in {'Eligible', 'Eligible (Bypass)'}:
                continue
            eligible_students.append(sid)
            has_any_session_content = bool(advised or optional or repeat or str(note).strip())
            if course_code in advised:
                advised_students.append(sid)
                if course_code in optional:
                    optional_advised_students.append(sid)
            elif has_any_session_content:
                not_advised_students.append(sid)
            else:
                skipped_advising_students.append(sid)
            if is_graduating:
                if has_any_session_content:
                    attended_graduating.append(sid)
                else:
                    skipped_graduating.append(sid)
        qaa_data.append({
            'course_code': str(course_code),
            'course_name': course_name,
            'eligibility': len(eligible_students),
            'advised': len(advised_students),
            'optional': len(optional_advised_students),
            'not_advised': len(not_advised_students),
            'skipped_advising': len(skipped_advising_students),
            'attended_graduating': len(attended_graduating),
            'skipped_graduating': len(skipped_graduating),
        })
    return sorted(qaa_data, key=lambda row: (row['eligibility'], row['advised']), reverse=True)


def _build_schedule_combinations(selections: Iterable[StudentSelection]) -> tuple[list[dict[str, Any]], int]:
    combo_counts: dict[tuple[str, ...], list[str]] = {}
    students_processed = 0
    for selection in selections:
        advised_only = sorted([course for course in selection.advised if course not in selection.optional])
        if len(advised_only) < 2:
            continue
        students_processed += 1
        key = tuple(advised_only)
        combo_counts.setdefault(key, []).append(str(selection.student_id))
    combo_data = []
    for courses, student_ids in combo_counts.items():
        combo_data.append({
            'courses': list(courses),
            'students': student_ids,
            '# Students': len(student_ids),
            '# Courses': len(courses),
            'Courses': ', '.join(courses),
        })
    combo_data.sort(key=lambda item: (item['# Students'], item['# Courses']), reverse=True)
    return combo_data, students_processed


def _merge_schedule_groups(combo_data: list[dict[str, Any]], target_count: int, max_courses: int) -> list[dict[str, Any]]:
    groups = [
        {
            'courses': set(item['courses']),
            'students': set(item['students']),
        }
        for item in combo_data
    ]
    changed = True
    while changed and len(groups) > target_count:
        changed = False
        best_pair: tuple[int, int] | None = None
        best_overlap = 0
        for idx, left in enumerate(groups):
            for jdx in range(idx + 1, len(groups)):
                right = groups[jdx]
                merged_courses = left['courses'] | right['courses']
                if len(merged_courses) > max_courses:
                    continue
                overlap = len(left['courses'] & right['courses'])
                if overlap > best_overlap:
                    best_overlap = overlap
                    best_pair = (idx, jdx)
        if best_pair is None or best_overlap == 0:
            break
        idx, jdx = best_pair
        groups[idx]['courses'] |= groups[jdx]['courses']
        groups[idx]['students'] |= groups[jdx]['students']
        groups.pop(jdx)
        changed = True
    merged = []
    for item in groups:
        courses = sorted(item['courses'])
        students = sorted(item['students'])
        merged.append({
            'courses': courses,
            'students': students,
            '# Students': len(students),
            '# Courses': len(courses),
            'Courses': ', '.join(courses),
        })
    merged.sort(key=lambda item: (item['# Students'], item['# Courses']), reverse=True)
    return merged


def schedule_conflicts(
    session: Session,
    major_code: str,
    target_groups: int | None = None,
    max_courses_per_group: int = 10,
    min_students: int = 1,
    min_courses: int = 2,
) -> list[dict[str, Any]]:
    major = _major_or_error(session, major_code)
    period = current_period(session, major_code)
    if not period:
        return []
    selections = session.scalars(
        select(StudentSelection).where(StudentSelection.major_id == major.id, StudentSelection.period_id == period.id)
    ).all()
    combo_data, students_processed = _build_schedule_combinations(selections)
    merged_data = _merge_schedule_groups(combo_data, target_groups or max(1, min(10, len(combo_data) or 1)), max_courses_per_group) if combo_data else []
    filtered = [
        {
            'group_name': item['Courses'],
            'student_count': item['# Students'],
            'course_count': item['# Courses'],
            'courses': item['courses'],
            'student_ids': item['students'],
            'students_processed': students_processed,
        }
        for item in merged_data
        if item['# Students'] >= min_students and item['# Courses'] >= min_courses
    ]
    return filtered


def degree_plan_view(session: Session, major_code: str, student_id: str) -> dict[str, Any]:
    courses_df = _courses_df(session, major_code)
    progress_df = _progress_df(session, major_code)
    if courses_df.empty:
        raise ValueError('No course data available')
    if progress_df.empty:
        return {'student': None, 'legend': [], 'years': []}
    students_df = progress_df.copy()
    student_match = students_df.loc[students_df['ID'].astype(str) == str(student_id)]
    if student_match.empty:
        raise ValueError(f'Student not found: {student_id}')
    student_data = student_match.iloc[0]
    semesters = _get_semester_structure(courses_df)
    course_statuses = _get_student_course_statuses(student_data, courses_df, major_code, session)
    years = []
    for year_name, semester_list in sorted(_group_semesters_by_year(semesters).items()):
        year_entry = {'year_name': year_name, 'semesters': []}
        for semester_key in semester_list:
            semester_courses = semesters.get(semester_key, [])
            year_entry['semesters'].append({
                'semester_key': semester_key,
                'total_credits': sum(float(course['credits'] or 0) for course in semester_courses),
                'courses': [
                    {
                        **course,
                        'status': course_statuses.get(str(course['code']), 'not_eligible'),
                    }
                    for course in semester_courses
                ],
            })
        years.append(year_entry)
    return {
        'student': {
            'student_id': str(student_data.get('ID', '')),
            'student_name': str(student_data.get('NAME', '')),
            'standing': str(student_data.get('Standing', get_student_standing(float(student_data.get('# of Credits Completed', 0) or 0) + float(student_data.get('# Registered', 0) or 0)))),
            'remaining_credits': float(student_data.get('Remaining Credits', student_data.get('# Remaining', 0)) or 0),
        },
        'legend': [
            {'status': 'completed', 'label': 'Completed', 'icon': '🟢'},
            {'status': 'registered', 'label': 'Registered', 'icon': '🟡'},
            {'status': 'available', 'label': 'Available', 'icon': '🔵'},
            {'status': 'advised', 'label': 'Advised', 'icon': '🟠'},
            {'status': 'not_eligible', 'label': 'Not Eligible', 'icon': '⚪'},
            {'status': 'failed', 'label': 'Failed', 'icon': '🔴'},
        ],
        'years': years,
    }


def _get_semester_structure(courses_df: pd.DataFrame) -> dict[str, list[dict[str, Any]]]:
    semester_col = None
    for col in courses_df.columns:
        if 'suggested' in str(col).lower() and 'semester' in str(col).lower():
            semester_col = col
            break
    if not semester_col:
        return {}
    semesters: dict[str, list[dict[str, Any]]] = {}
    for _, course_row in courses_df.iterrows():
        semester_value = str(course_row.get(semester_col, '')).strip()
        if not semester_value or semester_value.lower() in {'nan', 'none'}:
            continue
        if '-' not in semester_value:
            continue
        parts = semester_value.split('-')
        if len(parts) != 2:
            continue
        semester_name = parts[0].strip()
        year_num = parts[1].strip()
        semester_key = f'{semester_name}-{year_num}'
        semesters.setdefault(semester_key, []).append({
            'code': str(course_row.get('Course Code', '')),
            'title': str(course_row.get('Course Title', course_row.get('Title', ''))),
            'credits': float(course_row.get('Credits', 0) or 0),
            'semester': semester_name,
            'year': year_num,
        })
    return semesters


def _group_semesters_by_year(semesters: dict[str, list[dict[str, Any]]]) -> dict[str, list[str]]:
    year_groups: dict[str, list[str]] = {}
    semester_order = {'fall': 0, 'spring': 1, 'summer': 2}
    for semester_key in semesters.keys():
        if '-' not in semester_key:
            continue
        _, year_num = semester_key.split('-', 1)
        year_name = f'Year {year_num.strip()}'
        year_groups.setdefault(year_name, []).append(semester_key)
    for year_name in year_groups:
        year_groups[year_name].sort(key=lambda item: semester_order.get(item.split('-')[0].lower(), 99))
    return year_groups


def _get_student_course_statuses(student_data: pd.Series, courses_df: pd.DataFrame, major_code: str, session: Session) -> dict[str, str]:
    statuses: dict[str, str] = {}
    mutual_pairs = get_mutual_concurrent_pairs(courses_df)
    major = _major_or_error(session, major_code)
    period = current_period(session, major_code)
    selections = _selection_map(session, major, period.id if period else None)
    slot = selections.get(str(student_data.get('ID')))
    advised = slot.advised if slot else []
    repeat = slot.repeat if slot else []
    optional = slot.optional if slot else []
    for _, course_row in courses_df.iterrows():
        course_code = str(course_row.get('Course Code', ''))
        student_status = str(student_data.get(course_code, '') or '').strip().lower()
        if course_code in repeat:
            statuses[course_code] = 'advised_repeat'
        elif course_code in advised or course_code in optional:
            statuses[course_code] = 'advised'
        elif student_status in {'c', 'completed', 'pass', 'p', 'taken'}:
            statuses[course_code] = 'completed'
        elif student_status in {'r', 'registered', 'current', 'cr', 'reg'}:
            statuses[course_code] = 'registered'
        elif student_status in {'f', 'fail', 'failed'}:
            statuses[course_code] = 'failed'
        else:
            status, _ = check_eligibility(student_data, course_code, [], courses_df, registered_courses=[], ignore_offered=True, mutual_pairs=mutual_pairs)
            statuses[course_code] = 'available' if status == 'Eligible' else 'not_eligible'
    return statuses


def build_individual_report(session: Session, major_code: str, student_id: str, selected_courses: list[str] | None = None) -> tuple[str, bytes]:
    payload = individual_student_view(session, major_code, student_id, selected_courses)
    data = {'ID': [payload['student_id']], 'NAME': [payload['student_name']]}
    for course in payload['selected_courses']:
        data[course] = [payload['statuses'][course]]
    df = pd.DataFrame(data)
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Student')
    apply_individual_compact_formatting(output=output, sheet_name='Student', course_cols=payload['selected_courses'])
    return f'Student_{student_id}.xlsx', output.getvalue()


def build_all_advised_report(session: Session, major_code: str) -> tuple[str, bytes]:
    courses_df = _courses_df(session, major_code)
    progress_df = _progress_df(session, major_code)
    major = _major_or_error(session, major_code)
    period = current_period(session, major_code)
    selections = _selection_map(session, major, period.id if period else None)
    if courses_df.empty or progress_df.empty:
        raise ValueError('Courses and progress datasets must be uploaded first')
    all_sel = [(sid, sel) for sid, sel in selections.items() if sel.advised]
    output = BytesIO()
    mutual_pairs = get_mutual_concurrent_pairs(courses_df)
    bypasses = _bypass_map(session, major)
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        index_data = []
        for sid, sel in all_sel:
            srow_match = progress_df.loc[progress_df['ID'].astype(str) == str(sid)]
            if srow_match.empty:
                continue
            srow = srow_match.iloc[0]
            data_rows = {'Course Code': [], 'Action': [], 'Eligibility Status': [], 'Justification': [], 'Bypass': []}
            student_bypasses = {cc: info for (student_key, cc), info in bypasses.items() if student_key == str(sid)}
            for cc in courses_df['Course Code'].dropna().tolist():
                status, just = check_eligibility(srow, cc, sel.advised, courses_df, registered_courses=[], mutual_pairs=mutual_pairs, bypass_map=student_bypasses)
                bypass_note = ''
                if check_course_completed(srow, cc):
                    action = 'Completed'
                    status = 'Completed'
                elif check_course_registered(srow, cc):
                    action = 'Registered'
                elif cc in sel.advised:
                    action = 'Advised'
                elif status == 'Eligible (Bypass)':
                    action = 'Eligible (Bypass)'
                    bypass_info = student_bypasses.get(cc, {})
                    bypass_note = bypass_info.get('note', '')
                else:
                    action = 'Eligible not chosen' if status == 'Eligible' else 'Not Eligible'
                data_rows['Course Code'].append(cc)
                data_rows['Action'].append(action)
                data_rows['Eligibility Status'].append(status)
                data_rows['Justification'].append(just)
                data_rows['Bypass'].append(bypass_note)
            sheet_name = str(sid)[:31]
            pd.DataFrame(data_rows).to_excel(writer, index=False, sheet_name=sheet_name)
            ws = writer.sheets[sheet_name]
            ws.insert_rows(1, 4)
            ws['A1'] = 'Advising Session Snapshot'
            ws['A1'].font = TITLE_FONT
            ws['A2'] = f'Student: {srow.get("NAME", sid)} ({sid})'
            ws['A3'] = f'Period: {period.period_code if period else "N/A"}'
            ws['A4'] = f'Generated: {datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")}'
            ws['A4'].font = SUBTITLE_FONT
            _style_header_row(ws, 5)
            _apply_action_column_colors(ws, header_row=5)
            ws.freeze_panes = 'A6'
            _autosize_columns(ws)

            advised_credits = 0
            optional_credits = 0
            for cc in sel.advised:
                course_info = courses_df.loc[courses_df['Course Code'] == cc]
                if not course_info.empty:
                    credits = float(course_info.iloc[0].get('Credits', 0) or 0)
                    advised_credits += credits
                    if cc in sel.optional:
                        optional_credits += credits
            index_data.append({'ID': sid, 'NAME': srow.get('NAME', ''), 'Credits Advised': int(advised_credits), 'Optional Credits': int(optional_credits)})
        pd.DataFrame(index_data).to_excel(writer, index=False, sheet_name='Index')
        index_ws = writer.sheets['Index']
        index_ws.insert_rows(1, 1)
        index_ws['A1'] = 'All Advised Students - Index'
        index_ws['A1'].font = TITLE_FONT
        _style_header_row(index_ws, 2)
        index_ws.freeze_panes = 'A3'
        _autosize_columns(index_ws)
    return 'All_Advised_Students.xlsx', output.getvalue()


def build_qaa_report(session: Session, major_code: str, graduating_threshold: int = 36) -> tuple[str, bytes]:
    qaa_df = pd.DataFrame(qaa_sheet(session, major_code, graduating_threshold))
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        qaa_df.to_excel(writer, index=False, sheet_name='QAA Sheet')
        ws = writer.sheets['QAA Sheet']
        ws.insert_rows(1, 2)
        ws['A1'] = f'QAA Sheet - {major_code}'
        ws['A1'].font = TITLE_FONT
        ws['A2'] = f'Generated: {datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")} | Graduating threshold: {graduating_threshold}'
        ws['A2'].font = SUBTITLE_FONT
        _style_header_row(ws, 3)
        ws.freeze_panes = 'A4'

        skipped_col = None
        for idx, cell in enumerate(ws[3], start=1):
            if str(cell.value or '').strip().lower() == 'skipped_graduating':
                skipped_col = idx
                break
        if skipped_col is not None:
            for row in range(4, ws.max_row + 1):
                value = ws.cell(row=row, column=skipped_col).value
                if isinstance(value, (int, float)) and value > 0:
                    for col in range(1, ws.max_column + 1):
                        ws.cell(row=row, column=col).fill = PatternFill(start_color='FFF4CC', end_color='FFF4CC', fill_type='solid')

        _autosize_columns(ws)
    return f'QAA_Sheet_{major_code}.xlsx', output.getvalue()


def build_schedule_conflicts_report(session: Session, major_code: str, target_groups: int | None = None, max_courses_per_group: int = 10, min_students: int = 1, min_courses: int = 2) -> tuple[str, bytes]:
    rows = schedule_conflicts(session, major_code, target_groups=target_groups, max_courses_per_group=max_courses_per_group, min_students=min_students, min_courses=min_courses)
    df = pd.DataFrame(rows)
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Schedule Conflicts')
        ws = writer.sheets['Schedule Conflicts']
        _style_header_row(ws, 1)
        ws.freeze_panes = 'A2'
        _autosize_columns(ws)
    return f'schedule_conflict_{major_code}.xlsx', output.getvalue()
