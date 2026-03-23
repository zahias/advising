"""
Progress report processing service.
Ported from course_mapping/data_processing.py, config.py, and completion_utils.py.
All Streamlit dependencies removed; errors are raised as ValueError.
"""
from __future__ import annotations

import io
from typing import Any, Optional

import pandas as pd
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models import CourseAssignment, CourseEquivalent, CourseRule, DatasetVersion, Major

# ---------------------------------------------------------------------------
# Constants (from course_mapping/config.py)
# ---------------------------------------------------------------------------

GRADE_ORDER = [
    "CR",
    "A+", "A", "A-",
    "B+", "B", "B-",
    "C+", "C", "C-",
    "D+", "D", "D-",
]

COMPLETION_COLOR_MAP = {
    "c":  "#28a745",   # green  – completed
    "cr": "#FFFACD",   # yellow – currently registered
    "nc": "#f8d7da",   # red    – not completed
}


# ---------------------------------------------------------------------------
# Grade helpers (from course_mapping/config.py)
# ---------------------------------------------------------------------------

def is_passing_grade(grade: str, passing_grades_str: str) -> bool:
    try:
        passing = [g.strip().upper() for g in passing_grades_str.split(",")]
    except Exception:
        passing = []
    return grade.strip().upper() in passing


def extract_primary_grade_from_full_value(value: str) -> str:
    """Return the highest-priority single entry from a comma-joined 'GRADE | credit' string."""
    if not isinstance(value, str):
        return value
    entries = [e.strip() for e in value.split(",") if e.strip()]
    parsed = []
    for entry in entries:
        if "|" in entry:
            g, c = entry.split("|", 1)
            credit = c.strip()
            parsed.append({"grade": g.strip().upper(), "credit": credit, "credit_upper": credit.upper(), "original": entry})
    if not parsed:
        return value
    # CR has top priority; then highest GRADE_ORDER rank; then PASS token
    def rank(p: dict) -> int:
        if p["grade"] == "CR":
            return -1
        try:
            return GRADE_ORDER.index(p["grade"])
        except ValueError:
            return len(GRADE_ORDER)
    best = min(parsed, key=rank)
    return best["original"]


def cell_color(value: str) -> str:
    """Return a CSS background-color string for a processed grade cell value."""
    if not isinstance(value, str):
        return ""
    collapsed = value.strip().lower()
    color_css = {
        "c":  f"background-color: {COMPLETION_COLOR_MAP['c']}",
        "cr": f"background-color: {COMPLETION_COLOR_MAP['cr']}",
        "nc": f"background-color: {COMPLETION_COLOR_MAP['nc']}",
    }
    if collapsed in color_css:
        return color_css[collapsed]
    entries = [e.strip() for e in value.split(",") if e.strip()]
    for entry in entries:
        if entry.upper().startswith("CR"):
            return f"background-color: {COMPLETION_COLOR_MAP['cr']}"
    for entry in entries:
        parts = entry.split("|")
        if len(parts) == 2:
            right = parts[1].strip().upper()
            try:
                if int(right) > 0:
                    return f"background-color: {COMPLETION_COLOR_MAP['c']}"
            except ValueError:
                if right == "PASS":
                    return f"background-color: {COMPLETION_COLOR_MAP['c']}"
    return f"background-color: {COMPLETION_COLOR_MAP['nc']}"


# ---------------------------------------------------------------------------
# Completion utils (from course_mapping/completion_utils.py)
# ---------------------------------------------------------------------------

def collapse_pass_fail_value(val):
    """Normalize grade strings to completion shorthand: 'c', 'cr', or 'nc'."""
    if not isinstance(val, str):
        return val
    parts = [p.strip() for p in val.split("|")]
    if parts and parts[0].upper() == "CR":
        return "cr"
    if len(parts) == 1 and parts[0].upper() == "NR":
        return "nc"
    if len(parts) == 2:
        credit_str = parts[1]
        try:
            return "c" if int(credit_str) > 0 else "nc"
        except ValueError:
            normalized = credit_str.upper()
            if normalized == "PASS":
                return "c"
            if normalized == "FAIL":
                return "nc"
    return val


# ---------------------------------------------------------------------------
# Semester ordinal (from course_mapping/data_processing.py)
# ---------------------------------------------------------------------------

def semester_to_ordinal(semester: str, year) -> float:
    try:
        yr = int(year)
        sem = str(semester).strip().upper()
        sem_map = {"FALL": 0, "SPRING": 1, "SUMMER": 2}
        return yr * 3 + sem_map.get(sem, 0)
    except (ValueError, TypeError):
        return float('-inf')


# ---------------------------------------------------------------------------
# Wide-format detection and transformation
# ---------------------------------------------------------------------------

def _detect_file_format(df: pd.DataFrame) -> str:
    if {'Course', 'Grade', 'Year', 'Semester'}.issubset(df.columns):
        return "Long format (columns: Course, Grade, Year, Semester)"
    elif any(c.upper().startswith('COURSE') for c in df.columns):
        return "Wide format (COURSE_* columns detected)"
    return "Unknown format"


def transform_wide_format(df: pd.DataFrame) -> pd.DataFrame:
    """
    Convert wide-format progress sheet (COURSE_1 … COURSE_N cells) to long format.
    Each cell contains 'COURSECODE/SEMESTER-YEAR/GRADE'.
    Returns DataFrame with columns ['ID', 'NAME', 'Course', 'Grade', 'Year', 'Semester'].
    Raises ValueError on malformed data.
    """
    if 'ID' in df.columns:
        id_col = 'ID'
    elif 'STUDENT ID' in df.columns:
        id_col = 'STUDENT ID'
    else:
        raise ValueError("Wide format file missing an 'ID' or 'STUDENT ID' column.")

    if 'NAME' not in df.columns and 'Name' in df.columns:
        df = df.rename(columns={'Name': 'NAME'})
    if 'NAME' not in df.columns:
        raise ValueError("Wide format file missing a 'NAME' column.")

    course_cols = [c for c in df.columns if c.upper().startswith('COURSE')]
    if not course_cols:
        raise ValueError("Wide format file missing any 'COURSE…' columns.")

    id_vars = [c for c in df.columns if c not in course_cols]
    df_melt = df.melt(id_vars=id_vars, value_vars=course_cols, var_name='Course_Column', value_name='CourseData')
    df_melt = df_melt[df_melt['CourseData'].notna() & (df_melt['CourseData'].str.strip() != '')]

    parts = df_melt['CourseData'].str.split('/', expand=True)
    if parts.shape[1] < 3:
        sample = df_melt['CourseData'].dropna().head(3).tolist()
        raise ValueError(
            f"Expected format 'CODE/SEM-YYYY/GRADE'. Sample data: {sample}. "
            "Each course cell must contain: COURSECODE/SEMESTER-YEAR/GRADE"
        )

    df_melt['Course'] = parts[0].str.strip().str.upper()
    df_melt['RawSemYear'] = parts[1].str.strip()
    df_melt['Grade'] = parts[2].str.strip().str.upper()

    sem_parts = df_melt['RawSemYear'].str.split('-', expand=True)
    if sem_parts.shape[1] < 2:
        sample = df_melt['RawSemYear'].dropna().head(3).tolist()
        raise ValueError(
            f"Expected Semester-Year format 'SEMESTER-YYYY'. Sample: {sample}. "
            "Examples: Fall-2022, Spring-2023"
        )

    df_melt['Semester'] = sem_parts[0].str.strip().str.title()
    df_melt['Year'] = sem_parts[1].str.strip()

    if id_col != 'ID':
        df_melt = df_melt.rename(columns={id_col: 'ID'})

    final_cols = ['ID', 'NAME', 'Course', 'Grade', 'Year', 'Semester']
    missing = [c for c in final_cols if c not in df_melt.columns]
    if missing:
        raise ValueError(f"Missing columns after transformation: {missing}")

    return df_melt[final_cols].drop_duplicates()


# ---------------------------------------------------------------------------
# Core processing logic (from course_mapping/data_processing.py)
# ---------------------------------------------------------------------------

def determine_course_value(grade: str, course: str, courses_dict: dict, rules_list: list, term_ord: float = None) -> str:
    """
    Convert a raw grade string + course rules into a processed cell value
    like 'B+ | 3', 'CR | 3', 'NR', 'W | FAIL', etc.
    """
    credits = 0
    passing = ""

    if rules_list:
        matched_rule = None
        if term_ord is not None:
            for rule in rules_list:
                from_ord = rule.get("FromOrd", float('-inf'))
                to_ord = rule.get("ToOrd", float('inf'))
                if from_ord <= term_ord <= to_ord:
                    matched_rule = rule
                    break
        if matched_rule is None:
            matched_rule = rules_list[0]
        credits = matched_rule["Credits"]
        passing = matched_rule["PassingGrades"]

    if pd.isna(grade):
        return "NR"
    elif grade == "":
        return f"CR | {credits}" if credits > 0 else "CR | PASS"
    else:
        tokens = [g.strip().upper() for g in grade.split(", ") if g.strip()]
        all_toks = ", ".join(tokens)
        allowed = [x.strip().upper() for x in passing.split(",")] if passing else []
        # No restriction (empty allowed) means all grades pass
        passed = True if not allowed else any(g in allowed for g in tokens)
        if credits > 0:
            return f"{all_toks} | {credits}" if passed else f"{all_toks} | 0"
        else:
            return f"{all_toks} | PASS" if passed else f"{all_toks} | FAIL"


def process_progress_report(
    df: pd.DataFrame,
    target_courses: dict,
    intensive_courses: dict,
    target_rules: dict,
    intensive_rules: dict,
    assignment_types: list[str],
    per_student_assignments: Optional[dict] = None,
    equivalent_courses_mapping: Optional[dict] = None,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, list]:
    """
    Process long-format progress data into pivoted required / intensive tables.

    Returns (required_df, intensive_df, extra_courses_df, extra_courses_list).
    """
    if equivalent_courses_mapping is None:
        equivalent_courses_mapping = {}

    df = df.copy()

    # 1) Map equivalents
    df["Mapped Course"] = df["Course"].apply(lambda x: equivalent_courses_mapping.get(x, x))

    # 2) Apply assignment types (S.C.E., F.E.C., etc.)
    if per_student_assignments:
        def map_assignment(row):
            sid = str(row["ID"])
            course = row["Course"]
            if sid in per_student_assignments:
                assigns = per_student_assignments[sid]
                for atype in assignment_types:
                    if assigns.get(atype) == course:
                        return atype
            return row["Mapped Course"]
        df["Mapped Course"] = df.apply(map_assignment, axis=1)

    # 3) Term ordinal
    df["TermOrd"] = df.apply(
        lambda r: semester_to_ordinal(r.get("Semester", ""), r.get("Year", "")),
        axis=1,
    )

    # 4) ProcessedValue per row
    df["ProcessedValue"] = df.apply(
        lambda r: determine_course_value(
            r["Grade"],
            r["Mapped Course"],
            target_courses if r["Mapped Course"] in target_courses
            else intensive_courses if r["Mapped Course"] in intensive_courses
            else {},
            (target_rules.get(r["Mapped Course"], [])
             if r["Mapped Course"] in target_rules
             else intensive_rules.get(r["Mapped Course"], [])),
            r["TermOrd"],
        ),
        axis=1,
    )

    roster_df = df[["ID", "NAME"]].drop_duplicates()

    # 5) Split
    extra_courses_df = df[
        (~df["Mapped Course"].isin(target_courses)) &
        (~df["Mapped Course"].isin(intensive_courses))
    ]
    target_df = df[df["Mapped Course"].isin(target_courses)]
    intensive_df = df[df["Mapped Course"].isin(intensive_courses)]

    # 6) Pivot
    def _pivot(subset: pd.DataFrame) -> pd.DataFrame:
        if subset.empty:
            return roster_df.copy()
        piv = subset.pivot_table(
            index=["ID", "NAME"],
            columns="Mapped Course",
            values="ProcessedValue",
            aggfunc=lambda vals: ", ".join(vals),
        ).reset_index()
        return roster_df.merge(piv, on=["ID", "NAME"], how="left")

    pivot_df = _pivot(target_df)
    intensive_pivot_df = _pivot(intensive_df)

    # 7) Fill NR
    for course in target_courses:
        if course not in pivot_df.columns:
            pivot_df[course] = "NR"
        else:
            pivot_df[course] = pivot_df[course].fillna("NR")
    for course in intensive_courses:
        if course not in intensive_pivot_df.columns:
            intensive_pivot_df[course] = "NR"
        else:
            intensive_pivot_df[course] = intensive_pivot_df[course].fillna("NR")

    result_df = pivot_df[["ID", "NAME"] + list(target_courses.keys())]
    intensive_result_df = intensive_pivot_df[["ID", "NAME"] + list(intensive_courses.keys())]

    # 8) Remove assigned courses from extras
    if per_student_assignments:
        assigned_pairs = {
            (sid, crs)
            for sid, assigns in per_student_assignments.items()
            for crs in assigns.values()
        }
        extra_courses_df = extra_courses_df[
            ~extra_courses_df.apply(lambda row: (str(row["ID"]), row["Course"]) in assigned_pairs, axis=1)
        ]

    extra_courses_list = sorted(extra_courses_df["Course"].unique())
    return result_df, intensive_result_df, extra_courses_df, extra_courses_list


def calculate_credits(row: pd.Series, courses_dict: dict) -> pd.Series:
    """Tally completed / registered / remaining / total credits for a student row."""
    completed, registered, remaining = 0, 0, 0
    total = sum(courses_dict.values())

    for course, cred in courses_dict.items():
        val = row.get(course, "")
        if isinstance(val, str):
            entries = [e.strip() for e in val.split(",") if e.strip()]
            if any(e.upper().startswith("CR") for e in entries):
                registered += cred
                continue
            passed = False
            for e in entries:
                parts = [p.strip() for p in e.split("|")]
                if len(parts) == 2:
                    try:
                        if int(parts[1]) > 0:
                            passed = True
                            break
                    except ValueError:
                        if parts[1].upper() == "PASS":
                            passed = True
                            break
            if passed:
                completed += cred
            else:
                remaining += cred
        else:
            remaining += cred

    return pd.Series(
        [completed, registered, remaining, total],
        index=["# of Credits Completed", "# Registered", "# Remaining", "Total Credits"],
    )


# ---------------------------------------------------------------------------
# DB helpers – load rules/equivalents/assignments from DB
# ---------------------------------------------------------------------------

def _load_rules_from_db(session: Session, major_id: int) -> tuple[dict, dict]:
    """
    Load CourseRule rows from DB grouped by course_code.
    Returns (target_rules, intensive_rules) where each is:
      { course_code: [ {Credits, PassingGrades, FromOrd, ToOrd}, ... ] }
    We return a single combined dict; callers can separate by course type.
    """
    from app.models import CourseRule
    rows = session.scalars(select(CourseRule).where(CourseRule.major_id == major_id)).all()
    rules: dict[str, list] = {}
    for row in rows:
        from_ord = semester_to_ordinal(*row.from_semester.split("-", 1)) if row.from_semester else float('-inf')
        to_ord = semester_to_ordinal(*row.to_semester.split("-", 1)) if row.to_semester else float('inf')
        entry = {
            "Credits": row.credits,
            "PassingGrades": row.passing_grades or "",
            "FromOrd": from_ord,
            "ToOrd": to_ord,
            "course_type": row.course_type,
        }
        rules.setdefault(row.course_code.upper(), []).append(entry)
    return rules


def _load_equivalents_from_db(session: Session, major_id: int) -> dict[str, str]:
    rows = session.scalars(select(CourseEquivalent).where(CourseEquivalent.major_id == major_id)).all()
    return {row.alias_code.upper(): row.canonical_code.upper() for row in rows}


def _load_assignments_from_db(session: Session, major_id: int) -> dict[str, dict[str, str]]:
    rows = session.scalars(select(CourseAssignment).where(CourseAssignment.major_id == major_id)).all()
    result: dict[str, dict[str, str]] = {}
    for row in rows:
        result.setdefault(row.student_id, {})[row.assignment_type] = row.course_code.upper()
    return result


# ---------------------------------------------------------------------------
# Top-level orchestrator
# ---------------------------------------------------------------------------

def process_transcript(
    raw_df: pd.DataFrame,
    major_id: int,
    session: Session,
    courses_df: pd.DataFrame,
) -> bytes:
    """
    Full pipeline: raw wide-format DataFrame → two-sheet Excel bytes.

    courses_df must have columns: Course Code, Credits, Type
    (and optionally PassingGrades, RuleFromSemester, RuleToSemester — already
     loaded into DB by the dataset service; we use DB rules here).

    Returns bytes that load_progress_excel() can consume unchanged.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils.dataframe import dataframe_to_rows

    # 1) Transform wide → long
    long_df = transform_wide_format(raw_df)

    # 2) Build course lists from courses_df
    code_col = next((c for c in courses_df.columns if 'course' in c.lower() and 'code' in c.lower()), None)
    if code_col is None:
        code_col = courses_df.columns[0]
    cred_col = next((c for c in courses_df.columns if 'credit' in c.lower()), None)
    type_col = next((c for c in courses_df.columns if 'type' in c.lower()), None)

    courses_df = courses_df.copy()
    courses_df[code_col] = courses_df[code_col].astype(str).str.strip().str.upper()

    def _course_dict(type_filter: str) -> dict[str, int]:
        if type_col:
            subset = courses_df[courses_df[type_col].astype(str).str.strip().str.lower() == type_filter.lower()]
        else:
            subset = courses_df
        if cred_col:
            return {row[code_col]: int(row[cred_col]) if pd.notna(row[cred_col]) else 0 for _, row in subset.iterrows()}
        return {row[code_col]: 0 for _, row in subset.iterrows()}

    target_courses = _course_dict("required")
    intensive_courses = _course_dict("intensive")

    # 3) Load rules, equivalents, assignments from DB
    all_rules = _load_rules_from_db(session, major_id)
    target_rules = {k: v for k, v in all_rules.items() if k in target_courses}
    intensive_rules = {k: v for k, v in all_rules.items() if k in intensive_courses}
    equivalents = _load_equivalents_from_db(session, major_id)
    assignments = _load_assignments_from_db(session, major_id)

    # Load assignment_types from Major
    major = session.get(Major, major_id)
    assignment_types = (major.assignment_types or ["S.C.E", "F.E.C"]) if major else ["S.C.E", "F.E.C"]

    # 4) Process
    req_df, int_df, extra_df, _ = process_progress_report(
        long_df,
        target_courses,
        intensive_courses,
        target_rules,
        intensive_rules,
        assignment_types,
        per_student_assignments=assignments,
        equivalent_courses_mapping=equivalents,
    )

    # 5) Add credit columns
    req_df = req_df.join(req_df.apply(calculate_credits, axis=1, courses_dict=target_courses))
    int_df = int_df.join(int_df.apply(calculate_credits, axis=1, courses_dict=intensive_courses))

    # 6) Serialize to two-sheet Excel
    output = io.BytesIO()
    wb = Workbook()
    ws_req = wb.active
    ws_req.title = "Required Courses"

    completed_fill = PatternFill(start_color="28A745", end_color="28A745", fill_type="solid")
    current_fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")
    incomplete_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")

    def _write_sheet(ws, df: pd.DataFrame) -> None:
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    if value == "c":
                        cell.fill = completed_fill
                    elif value in ("nc", ""):
                        cell.fill = incomplete_fill
                    elif isinstance(value, str):
                        style = cell_color(value)
                        if COMPLETION_COLOR_MAP["c"] in style:
                            cell.fill = completed_fill
                        elif COMPLETION_COLOR_MAP["cr"] in style:
                            cell.fill = current_fill
                        else:
                            cell.fill = incomplete_fill

    _write_sheet(ws_req, req_df)
    ws_int = wb.create_sheet(title="Intensive Courses")
    _write_sheet(ws_int, int_df)

    # 3rd sheet: extra (unassigned) courses
    ws_extra = wb.create_sheet(title="Extra Courses")
    extra_cols = ["ID", "NAME", "Course", "Grade", "Year", "Semester"]
    available_extra_cols = [c for c in extra_cols if c in extra_df.columns]
    for c_idx, col in enumerate(available_extra_cols, 1):
        hdr = ws_extra.cell(row=1, column=c_idx, value=col)
        hdr.font = Font(bold=True)
        hdr.alignment = Alignment(horizontal="center")
    for r_idx, row_data in enumerate(
        extra_df[available_extra_cols].itertuples(index=False), 2
    ):
        for c_idx, val in enumerate(row_data, 1):
            ws_extra.cell(row=r_idx, column=c_idx, value=val)

    wb.save(output)
    output.seek(0)
    return output.read()
