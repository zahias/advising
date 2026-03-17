from __future__ import annotations

from fastapi import APIRouter, Depends
from fastapi.responses import Response
from sqlalchemy.orm import Session

from app.api.deps import ensure_major_access, get_db, require_staff
from app.models import User
from fastapi import Query

from app.services.insights_service import build_all_advised_report, build_individual_report, build_qaa_report, build_schedule_conflicts_report
from app.services.student_service import export_student_report

router = APIRouter(prefix='/reports', tags=['reports'])


@router.get('/{major_code}/student/{student_id}')
def student_report_route(major_code: str, student_id: str, user: User = Depends(require_staff), db: Session = Depends(get_db)):
    ensure_major_access(major_code, db, user)
    filename, payload = export_student_report(db, major_code, student_id)
    headers = {'Content-Disposition': f'attachment; filename={filename}'}
    return Response(content=payload, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers=headers)


@router.get('/{major_code}/individual/{student_id}')
def individual_compact_report_route(
    major_code: str,
    student_id: str,
    courses: list[str] = Query(default=[]),
    user: User = Depends(require_staff),
    db: Session = Depends(get_db),
):
    ensure_major_access(major_code, db, user)
    filename, payload = build_individual_report(db, major_code, student_id, courses or None)
    headers = {'Content-Disposition': f'attachment; filename={filename}'}
    return Response(content=payload, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers=headers)


@router.get('/{major_code}/all-advised')
def all_advised_report_route(major_code: str, user: User = Depends(require_staff), db: Session = Depends(get_db)):
    ensure_major_access(major_code, db, user)
    filename, payload = build_all_advised_report(db, major_code)
    headers = {'Content-Disposition': f'attachment; filename={filename}'}
    return Response(content=payload, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers=headers)


@router.get('/{major_code}/qaa')
def qaa_report_route(
    major_code: str,
    graduating_threshold: int = Query(default=36),
    user: User = Depends(require_staff),
    db: Session = Depends(get_db),
):
    ensure_major_access(major_code, db, user)
    filename, payload = build_qaa_report(db, major_code, graduating_threshold)
    headers = {'Content-Disposition': f'attachment; filename={filename}'}
    return Response(content=payload, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers=headers)


@router.get('/{major_code}/schedule-conflicts')
def schedule_conflict_report_route(
    major_code: str,
    target_groups: int = Query(default=10),
    max_courses_per_group: int = Query(default=10),
    min_students: int = Query(default=1),
    min_courses: int = Query(default=2),
    user: User = Depends(require_staff),
    db: Session = Depends(get_db),
):
    ensure_major_access(major_code, db, user)
    filename, payload = build_schedule_conflicts_report(
        db,
        major_code,
        target_groups=target_groups,
        max_courses_per_group=max_courses_per_group,
        min_students=min_students,
        min_courses=min_courses,
    )
    headers = {'Content-Disposition': f'attachment; filename={filename}'}
    return Response(content=payload, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers=headers)


@router.get('/{major_code}/progress-report')
def progress_report_export(major_code: str, user: User = Depends(require_staff), db: Session = Depends(get_db)):
    """Download a color-coded Excel progress report for all students."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils.dataframe import dataframe_to_rows
    import pandas as pd

    ensure_major_access(major_code, db, user)

    from app.services.dataset_service import get_active_dataset, load_progress_excel
    from app.services.storage import StorageService
    from app.services.progress_service import COMPLETION_COLOR_MAP, cell_color

    progress_v = get_active_dataset(db, major_code, 'progress')
    if not progress_v:
        from fastapi import HTTPException
        raise HTTPException(status_code=404, detail='No active progress dataset for this major.')

    storage = StorageService()
    file_bytes = storage.get_bytes(progress_v.storage_key)

    sheets = pd.read_excel(io.BytesIO(file_bytes), sheet_name=None)
    req_key = next((k for k in sheets if 'required' in k.lower()), None)
    int_key = next((k for k in sheets if 'intensive' in k.lower()), None)
    req_df = sheets.get(req_key, pd.DataFrame()) if req_key else pd.DataFrame()
    int_df = sheets.get(int_key, pd.DataFrame()) if int_key else pd.DataFrame()

    completed_fill = PatternFill(start_color='28A745', end_color='28A745', fill_type='solid')
    current_fill = PatternFill(start_color='FFFACD', end_color='FFFACD', fill_type='solid')
    incomplete_fill = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')

    def _write_sheet(ws, df: pd.DataFrame) -> None:
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif isinstance(value, str):
                    style = cell_color(value)
                    if COMPLETION_COLOR_MAP['c'] in style:
                        cell.fill = completed_fill
                    elif COMPLETION_COLOR_MAP['cr'] in style:
                        cell.fill = current_fill
                    else:
                        cell.fill = incomplete_fill

    wb = Workbook()
    ws_req = wb.active
    ws_req.title = 'Required Courses'
    _write_sheet(ws_req, req_df)
    if not int_df.empty:
        ws_int = wb.create_sheet(title='Intensive Courses')
        _write_sheet(ws_int, int_df)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    headers = {'Content-Disposition': f'attachment; filename="{major_code}_progress_report.xlsx"'}
    return Response(
        content=output.read(),
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers=headers,
    )
