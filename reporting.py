# reporting.py
# Excel formatting utilities for advising reports

from io import BytesIO
from typing import Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook


# Color mapping for course status codes
STATUS_COLORS = {
    "c": "C6E0B4",    # Completed -> light green
    "r": "BDD7EE",    # Registered -> light blue
    "s": "D9E1F2",    # Simulated (will register) -> purple/blue
    "a": "FFF2CC",    # Advised -> light yellow
    "ar": "FFD966",   # Advised-Repeat -> darker yellow/orange
    "o": "FFE699",    # Optional -> light orange
    "na": "E1F0FF",   # Eligible not chosen -> light blue-tint
    "ne": "F8CECC",   # Not Eligible -> light red
}


def apply_excel_formatting(
    output: BytesIO,
    student_name: str,
    student_id: int,
    credits_completed: int,
    standing: str,
    note: str,
    advised_credits: int,
    optional_credits: int,
    period_info: str = ""
):
    """
    Apply Excel formatting to individual student advising sheet.
    Adds header with student info and formats the table.
    """
    output.seek(0)
    wb = load_workbook(output)
    ws = wb.active
    
    ws.insert_rows(1, 7)
    
    ws["A1"] = "Student Advising Sheet"
    ws["A1"].font = Font(bold=True, size=16)
    
    if period_info:
        ws["A2"] = period_info
        ws["A2"].font = Font(bold=True, color="0066CC")
    
    ws["A3"] = f"Name: {student_name}"
    ws["A4"] = f"ID: {student_id}"
    ws["A5"] = f"Credits Completed: {credits_completed} | Standing: {standing}"
    ws["A6"] = f"Advised Credits: {advised_credits} | Optional Credits: {optional_credits}"
    
    if note:
        ws["A7"] = f"Notes: {note}"
        ws["A7"].font = Font(italic=True)
    
    header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[8]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    for row in ws.iter_rows(min_row=9):
        for cell in row:
            cell.alignment = Alignment(horizontal="left", vertical="center")
    
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    output.seek(0)
    wb.save(output)
    output.seek(0)


def add_summary_sheet(writer: pd.ExcelWriter, df: pd.DataFrame, course_cols: list):
    """
    Add a summary sheet with course statistics to the Excel workbook.
    """
    summary_data = []
    
    for course in course_cols:
        if course in df.columns:
            values = df[course].value_counts()
            summary_data.append({
                "Course": course,
                "Completed (c)": values.get("c", 0),
                "Registered (r)": values.get("r", 0),
                "Simulated (s)": values.get("s", 0),
                "Advised (a)": values.get("a", 0),
                "Advised-Repeat (ar)": values.get("ar", 0),
                "Optional (o)": values.get("o", 0),
                "Eligible Not Chosen (na)": values.get("na", 0),
                "Not Eligible (ne)": values.get("ne", 0),
            })
    
    summary_df = pd.DataFrame(summary_data)
    summary_df.to_excel(writer, index=False, sheet_name="Summary")


def _format_status_columns(ws, course_cols: Iterable[str]):
    header_row = next(ws.iter_rows(min_row=1, max_row=1))
    header_values = [cell.value for cell in header_row]

    course_col_indices: list[int] = []
    for course in course_cols:
        if course in header_values:
            course_col_indices.append(header_values.index(course) + 1)

    if not course_col_indices:
        return

    for row in ws.iter_rows(min_row=2):
        for col_idx in course_col_indices:
            cell = row[col_idx - 1]
            value = str(cell.value).strip().lower() if cell.value else ""
            if value in STATUS_COLORS:
                cell.fill = PatternFill(
                    start_color=STATUS_COLORS[value],
                    end_color=STATUS_COLORS[value],
                    fill_type="solid",
                )


def apply_full_report_formatting(output_or_workbook, sheet_name: str, course_cols: list):
    """
    Apply color formatting to full report with multiple students.
    Colors course status cells based on status code.
    """
    if not course_cols:
        return

    if isinstance(output_or_workbook, Workbook):
        wb = output_or_workbook
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            _format_status_columns(ws, course_cols)
        return

    output: BytesIO = output_or_workbook
    output.seek(0)
    wb = load_workbook(output)

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        _format_status_columns(ws, course_cols)

    output.seek(0)
    output.truncate(0)
    wb.save(output)
    wb.close()
    output.seek(0)


def apply_individual_compact_formatting(output: BytesIO, sheet_name: str, course_cols: list):
    """
    Apply color formatting to individual student compact report.
    Similar to full report formatting but for single student view.
    """
    apply_full_report_formatting(output, sheet_name, course_cols)
